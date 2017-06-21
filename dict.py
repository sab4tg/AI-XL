#written by Scott Bergstresser, Christian Okada, Cameron Mayes, and Brandon Bui
#simple machine learning exercise, takes in categorical data from a training sets
#   and constructs dictionaries that map specific data to specific categories.
#   Program then reads in a second excel file with similar data fields
#   and chooses the best category based on the given data, weighting the
#   categories based on their importance.
#Accenture internship Summer 2017

from __future__ import unicode_literals
from openpyxl import load_workbook


cd = {} # counter dictionary, used to determine the correct category
#dictionaries that map data to the 31 possible categories
ids,office,org,carl,fun,title,tt,tloc,cor,cname,curl,ccn,act= ({} for i in range(13))
charstr = ['B', 'D', 'F', 'I', 'J', 'L', 'M','P', 'Q', 'W', 'X', 'Z', 'AB']
dlist = []
dlist.extend((ids,office,org,carl,fun,title,tt,tloc,cor,cname,curl,ccn,act))



wb = load_workbook('contest-train.xlsx') #crack open a cold workbook with the boys
ws = wb['contest-train']
catlist=[] #populate a list of possible categories
for row in range(2,ws.max_row):
    if ws['AA' + str(row)].value not in catlist:
        catlist.append(ws['AA' + str(row)].value)
    if ws['AA' + str(row)].value not in cd.keys():
        cd[ws['AA' + str(row)].value] = 0

for cat in catlist:# initialize
    cd[cat] = 0

#initialize dictionaries, these map our data sets to categories. Each dictionary
#has keys representing the 31 categories, values are an array of data learned
#from the training set
for dic in dlist:
    for cat in catlist:
        dic[cat]=[]

#############################Populate the dictionaries####################################
for row in range(2,ws.max_row):
    if (ws['B'+str(row)].value is not 'NA') and (ws['B'+str(row)].value is not 'N/A'):
        if ws['B'+str(row)].value not in ids[ws['AA'+str(row)].value]:
            ids[ws['AA'+str(row)].value].append(ws['B'+str(row)].value.encode("ascii", "ignore"))

    if (ws['D'+str(row)].value is not 'NA') and (ws['D'+str(row)].value is not 'N/A'):
        if ws['D'+str(row)].value not in office[ws['AA'+str(row)].value]:
            office[ws['AA'+str(row)].value].append(ws['D'+str(row)].value.encode("ascii", "ignore"))

    if (ws['F'+str(row)].value is not 'NA') and (ws['F'+str(row)].value is not 'N/A'):
        if ws['F'+str(row)].value not in org[ws['AA'+str(row)].value]:
            org[ws['AA'+str(row)].value].append(ws['F'+str(row)].value.encode("ascii", "ignore"))

    if (ws['I'+str(row)].value is not 'NA') and (ws['I'+str(row)].value is not 'N/A'):
        if ws['I'+str(row)].value not in carl[ws['AA'+str(row)].value]:
            carl[ws['AA'+str(row)].value].append(ws['I'+str(row)].value)#.encode("ascii", "ignore"))

    if (ws['J'+str(row)].value is not 'NA') and (ws['J'+str(row)].value is not 'N/A'):
        if ws['J'+str(row)].value not in fun[ws['AA'+str(row)].value]:
            fun[ws['AA'+str(row)].value].append(ws['J'+str(row)].value.encode("ascii", "ignore"))

    if (ws['L'+str(row)].value is not 'NA') and (ws['L'+str(row)].value is not 'N/A'):
        if ws['L'+str(row)].value not in title[ws['AA'+str(row)].value]:
            title[ws['AA'+str(row)].value].append(ws['L'+str(row)].value.encode("ascii", "ignore"))

    if (ws['M'+str(row)].value is not 'NA') and (ws['M'+str(row)].value is not 'N/A'):
        if ws['M'+str(row)].value not in tt[ws['AA'+str(row)].value]:
            tt[ws['AA'+str(row)].value].append(ws['M'+str(row)].value.encode("ascii", "ignore"))

    if (ws['P'+str(row)].value is not 'NA') and (ws['P'+str(row)].value is not 'N/A'):
        if ws['P'+str(row)].value not in tloc[ws['AA'+str(row)].value]:
            tloc[ws['AA'+str(row)].value].append(ws['P'+str(row)].value)#.encode("ascii", "ignore"))

    if (ws['Q'+str(row)].value is not 'NA') and (ws['Q'+str(row)].value is not 'N/A'):
        if ws['Q'+str(row)].value not in cor[ws['AA'+str(row)].value]:
            cor[ws['AA'+str(row)].value].append(ws['Q'+str(row)].value.encode("ascii", "ignore"))

    if (ws['W'+str(row)].value is not 'NA') and (ws['W'+str(row)].value is not 'N/A'):
        if ws['W'+str(row)].value not in cname[ws['AA'+str(row)].value]:
            cname[ws['AA'+str(row)].value].append(ws['W'+str(row)].value)#.encode("ascii", "ignore"))

    if (ws['X'+str(row)].value is not 'NA') and (ws['X'+str(row)].value is not 'N/A'):
        if ws['X'+str(row)].value not in curl[ws['AA'+str(row)].value]:
            curl[ws['AA'+str(row)].value].append(ws['X'+str(row)].value)#.encode("ascii", "ignore"))

    if (ws['Z'+str(row)].value is not 'NA') and (ws['Z'+str(row)].value is not 'N/A'):
        if ws['Z'+str(row)].value not in ccn[ws['AA'+str(row)].value]:
            ccn[ws['AA'+str(row)].value].append(ws['Z'+str(row)].value.encode("ascii", "ignore"))

    if (ws['AB'+str(row)].value is not 'NA') and (ws['AB'+str(row)].value is not 'N/A'):
        if ws['AB'+str(row)].value not in act[ws['AA'+str(row)].value]:
            act[ws['AA'+str(row)].value].append(ws['AB'+str(row)].value.encode("ascii", "ignore"))

#################Dictionaries have been created, now to parse test set###################

wb2 = load_workbook('contest-test.xlsx') #cold workbook number 2
ws2 = wb2['contest-test']

#Data have different weights based on importance to categories from training set
for row in range(2,ws2.max_row):
    for cat in catlist:
        if ws2['D'+str(row)].value in office[cat]:
            cd[cat] += 1
        if ws2['F'+str(row)].value in org[cat]:
            cd[cat] += 2
        if ws2['I'+str(row)].value in carl[cat]:
            cd[cat] += 1
        if ws2['J'+str(row)].value in fun[cat]:
            cd[cat] += 1
        if ws2['L'+str(row)].value in title[cat]:
            cd[cat] += 5
        if ws2['M'+str(row)].value in tt[cat]:
            cd[cat] += 3
        if ws2['P'+str(row)].value in tloc[cat]:
            cd[cat] += 4
        if ws2['Q'+str(row)].value in cor[cat]:
            cd[cat] += 4
        if ws2['W'+str(row)].value in cname[cat]:
            cd[cat] += 3
        if ws2['X'+str(row)].value in curl[cat]:
            cd[cat] += 3
        if ws2['Z'+str(row)].value in ccn[cat]:
            cd[cat] += 2
        if ws2['AB'+str(row)].value in act[cat]:
            cd[cat] += 6
    maximum = max(cd, key=cd.get)
    maxcount = cd[maximum]
    ws2['AA'+str(row)].value=maximum
    for key in cd.keys():
        cd[key] = 0

wb2.save('contest-test.xlsx')
