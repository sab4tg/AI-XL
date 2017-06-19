#AB Q A
from __future__ import unicode_literals
import openpyxl
from openpyxl import load_workbook

dic = {}
wb = load_workbook('contest-train.xlsx')

print 'sheet names are ' + str(wb.get_sheet_names())
sheet_ranges = wb['contest-train']
# print sheet_ranges['AA1'].value

for row in range(2,2704):
    #print sheet_ranges['AA' + str(row)].value
    if sheet_ranges['AA' + str(row)].value not in dic:#category
        dic[sheet_ranges['AA' + str(row)].value] = []
    else:
        if sheet_ranges['Q' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['Q' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['M' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['M' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['T' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['T' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['Z' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['Z' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['AB' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['AB' + str(row)].value.encode("ascii", "ignore"))
        #if sheet_ranges['W' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
        #    dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['W' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['S' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['S' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['F' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['F' + str(row)].value.encode("ascii", "ignore"))
        if sheet_ranges['G' + str(row)].value not in dic[sheet_ranges['AA' + str(row)].value]:
            dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['G' + str(row)].value.encode("ascii", "ignore"))


        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['M' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['T' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['Z' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['AB' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['W' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['S' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['F' + str(row)].value)
        # dic[sheet_ranges['AA' + str(row)].value].append(sheet_ranges['G' + str(row)].value)


# for item in dic.keys():
#     print item
#     print "\t" + str(dic[item]) + "\n"


wb2 = load_workbook('contest-test.xlsx')
