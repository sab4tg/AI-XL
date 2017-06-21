from __future__ import unicode_literals
import openpyxl

# wb = Workbook()
wb = openpyxl.load_workbook('contest-train.xlsx')

# dict = {'ACN/AFS Curriculum': 'CDP'}

#loop through rows and populate category(AA) value
ws = wb.active

AB = []
Q = []

for row in ws.iter_rows('AB{}:AB{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        AB.append(cell.value.encode("ascii", "ignore"))

for row in ws.iter_rows('Q{}:Q{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        Q.append(cell.value.encode("ascii", "ignore"))
c = 2
for row in ws.iter_rows('AA{}:AA{}'.format(2,ws.max_row)):
    for cell in row:
        # cell.value = 'ACN/AFS Curriculum'
        print str(c)+":     AB:"+AB[c]+"\tQ:"+Q[c]+"\tAA:"+str(cell.value)
        c += 1

        if c > 2701:
            exit()

# Save the file
# wb.save("contest-test.xlsx")
